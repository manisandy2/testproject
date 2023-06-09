import sys
import os
import openpyxl
import pandas as pd
import requests
from googleplaces import GooglePlaces, types
from openpyxl import load_workbook


# ===================================================================#
#                       CHANGE HERE                                 #
# ===================================================================#
keyword = "Poovika Mobiles"  # keyword to be searched
location = sys.argv[1]
result_type = types.TYPE_SCHOOL
sheet_name = location  # Name of the sheet in excel file, sheet_name=location is default

# ============= DO NOT MODIFY ANYTHING BELOW THIS ====================
# API_KEY = 'YOUR API KEY HERE'  # your API key
API_KEY = "AIzaSyAYI0RbWinikf9Si0eg6oEfknDVwnRhXzY"

radius = 20000  # radius in meters
google_places = GooglePlaces(API_KEY)

names_list = list()
websites_list = list()
phone_numbers_list = list()

i = 0
next_page_token = ''
for i in range(0, 3):
    if next_page_token == '':
        query_result = google_places.nearby_search(
            location=location, keyword=keyword,
            radius=radius, types=[result_type])
    else:
        query_result = google_places.nearby_search(
            location=location, keyword=keyword,
            radius=radius, types=[result_type], pagetoken=next_page_token)

    # for each result of nearby search, extract place ID and get place details(which will contain the website)
    for result in query_result.raw_response['results']:
        parameters = {"key": API_KEY, "placeid": result['place_id']}  # extracting place ID
        response = requests.get("https://maps.googleapis.com/maps/api/place/details/json",
                                params=parameters)  # getting place details
        try:
            name = website = phone_number = ''
            if 'name' in response.json()['result'].keys():
                name = response.json()['result']['name']
            if 'website' in response.json()['result'].keys():
                website = '=HYPERLINK("' + response.json()['result']['website'] + '")'
            if 'formatted_phone_number' in response.json()['result'].keys():
                phone_number = response.json()['result']['formatted_phone_number']
        finally:
            names_list.append(name)
            websites_list.append(website)
            phone_numbers_list.append(phone_number)
            # time.sleep(1)  # time delay between consecutive API calls

    if 'next_page_token' in query_result.raw_response.keys():
        next_page_token = query_result.raw_response['next_page_token']
    else:
        next_page_token = ''
    i = i + 1

print(names_list)
print(websites_list)
print(phone_numbers_list)

dataframe = pd.DataFrame()  # create pandas data frame
dataframe['Name'] = names_list
dataframe['Website'] = websites_list
dataframe['Phone'] = phone_numbers_list

# creating excel file if does not exist
if not os.path.isfile('data.xlsx'):
    print("FILE NOT FOUND...CREATING ONE")
    wb = openpyxl.Workbook()
    wb.save('data.xlsx')

# Writing to excel sheet
excel_writer = pd.ExcelWriter("data.xlsx")  # write to excel sheet
book = load_workbook("data.xlsx")
excel_writer.book = book
dataframe.to_excel(excel_writer, sheet_name=location, index=False)
excel_writer.save()
excel_writer.close()

# Removing the default sheet if it exists
wb = load_workbook('data.xlsx')
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
wb.save('data.xlsx')